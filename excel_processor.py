# Copyright (c) Opendatalab. All rights reserved.
"""
Excel 文件独立处理器。

直接从 .xlsx/.xls 文件读取所有 Sheet, 转换为结构化 Markdown。
支持 Dify 表格优化, 支持超大文件 (>100MB) 的流式读取。

处理流程:
1. 检测文件大小 → 选择读取模式 (normal / read_only)
2. 遍历所有 Sheet → 转为 HTML 表格
3. 可选: 送 Dify 优化表格结构
4. 合并输出为 Markdown (每 Sheet 一个 ## 章节)

超大文件策略:
- < 50MB: normal 模式 (支持合并单元格)
- ≥ 50MB: read_only 流式模式 (每行迭代, 内存友好)
- 单 Sheet > 1000 行: 分块送 Dify (每 500 行一块)
- 单 Sheet > 5000 行: 跳过 Dify, 直接输出 Markdown (可配置)
"""
import asyncio
import os
import time
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Optional

from loguru import logger

# ── 配置常量 ──────────────────────────────────────────────
MAX_FILE_SIZE_NORMAL_MODE = 50 * 1024 * 1024    # 50MB — 超过此值用 read_only
DIFY_MAX_ROWS_PER_SHEET = 1000                   # 单 Sheet 超过此行数分批送 Dify
DIFY_SKIP_ROWS_THRESHOLD = 5000                  # 单 Sheet 超过此行数跳过 Dify
MAX_ROWS_PER_CHUNK = 500                         # Dify 分块大小
MAX_COLS_FOR_MARKDOWN_TABLE = 30                 # 超过此列数降级为 HTML 输出


@dataclass
class SheetResult:
    """单个 Sheet 的处理结果"""
    name: str
    row_count: int
    col_count: int
    html: str              # 原始 HTML 表格
    optimized_html: str    # Dify 优化后的 HTML (空=未优化)
    markdown: str          # 最终 Markdown 输出
    dify_enhanced: bool = False
    error: str = ""


@dataclass
class ExcelResult:
    """Excel 文件处理结果"""
    file_name: str
    sheets: list[SheetResult] = field(default_factory=list)
    total_rows: int = 0
    dify_calls: int = 0
    processing_time_s: float = 0.0
    errors: list[str] = field(default_factory=list)


# ── Excel 读取 ────────────────────────────────────────────

def _read_sheet_normal(ws) -> tuple[list[list], list[tuple]]:
    """
    Normal 模式读取 (支持合并单元格检测)。

    Returns:
        (rows_data, merged_cells_info)
        rows_data: [[cell_value, ...], ...]
        merged_cells_info: [(min_row, min_col, max_row, max_col), ...]
    """
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=False):
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        # 跳过完全空行
        if any(v is not None for v in row_data):
            rows.append(row_data)

    # 提取合并单元格信息
    merged = []
    for merge_range in ws.merged_cells.ranges:
        merged.append((
            merge_range.min_row - 1,   # 0-based
            merge_range.min_col - 1,
            merge_range.max_row - 1,
            merge_range.max_col - 1,
        ))

    return rows, merged


def _read_sheet_streaming(ws) -> list[list]:
    """
    Read-only 流式模式读取 (内存友好, 不支持合并单元格)。

    逐行迭代, 不缓存前面的行。
    """
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            rows.append(list(row))
    return rows


def _detect_excel_header_rows(rows: list[list], start_row: int = 0) -> int:
    """
    检测 Excel 表格中表头占用的行数。

    启发式规则:
    - 第 0 行始终是表头
    - 第 1 行有合并单元格 (colspan > 1) → 父级表头, 也是表头
    - 第 1 行文本平均长度 < 数据行平均的 50% → 子级表头
    - 最多 3 行表头
    """
    total = len(rows)
    if total <= 1:
        return 1 if total == 1 else 0

    # 计算数据行的平均文本长度 (跳过前 2 行)
    data_start = min(2, total)
    data_lengths = []
    for r in range(data_start, total):
        for val in (rows[r] or []):
            if val is not None:
                data_lengths.append(len(str(val)))
    avg_data = sum(data_lengths) / max(len(data_lengths), 1)

    header_count = 1  # row 0 always

    if total >= 3:
        row1 = rows[1] if len(rows) > 1 else []
        row1_lengths = [len(str(v)) for v in row1 if v is not None]
        avg_row1 = sum(row1_lengths) / max(len(row1_lengths), 1)

        # 检测是否有合并单元格 (在行切片范围内无法检测 merged_cells, 仅用文本)
        if avg_data > 0 and avg_row1 < avg_data * 0.5:
            header_count = 2

    # Row 2: 仅在大表 (> 10 行) 且 row1 已是表头时检测
    if header_count >= 2 and total >= 10:
        row2 = rows[2] if len(rows) > 2 else []
        row2_lengths = [len(str(v)) for v in row2 if v is not None]
        avg_row2 = sum(row2_lengths) / max(len(row2_lengths), 1)
        if avg_data > 0 and avg_row2 < avg_data * 0.4:
            header_count = 3

    return min(header_count, 3)


def _rows_to_html(
    rows: list[list],
    merged_cells: list[tuple] = None,
    max_rows: int = 0,
    start_row: int = 0,
) -> str:
    """
    将二维行数据转为 HTML 表格 (支持多层表头)。

    Args:
        rows: [[cell_value, ...], ...]
        merged_cells: [(min_r, min_c, max_r, max_c), ...] 0-based
        max_rows: 0 = 全部行
        start_row: 起始行偏移 (分块时用)

    Returns:
        HTML 字符串
    """
    if not rows:
        return ""

    col_count = max(len(r) for r in rows) if rows else 0
    if col_count == 0:
        return ""

    end = min(len(rows), start_row + max_rows) if max_rows > 0 else len(rows)
    visible = rows[start_row:end]

    # 构建合并单元格查找表
    merge_map = {}  # (r, c) → (rowspan, colspan)
    if merged_cells:
        for min_r, min_c, max_r, max_c in merged_cells:
            if start_row <= min_r < end:
                rs = max_r - min_r + 1
                cs = max_c - min_c + 1
                merge_map[(min_r, min_c)] = (rs, cs)

    # ★ 多层表头检测
    header_count = _detect_excel_header_rows(rows, start_row) if start_row == 0 else 1

    html = ["<table>"]

    # ── thead ──
    if header_count > 0:
        html.append("<thead>")
        for local_idx in range(min(header_count, len(visible))):
            html.append(_render_excel_row(
                visible[local_idx], col_count,
                start_row + local_idx, merge_map, "th",
            ))
        html.append("</thead>")

    # ── tbody ──
    html.append("<tbody>")
    for local_idx in range(header_count, len(visible)):
        html.append(_render_excel_row(
            visible[local_idx], col_count,
            start_row + local_idx, merge_map, "td",
        ))
    html.append("</tbody>")
    html.append("</table>")
    return "\n".join(html)


def _render_excel_row(
    row: list, col_count: int, global_row: int,
    merge_map: dict, tag: str,
) -> str:
    """渲染 Excel 表格的一行"""
    parts = ["<tr>"]
    col = 0
    while col < col_count:
        merge_info = merge_map.get((global_row, col))
        if merge_info:
            rs, cs = merge_info
            cell_val = row[col] if col < len(row) else ""
            content = _cell_to_text(cell_val)
            scope = ""
            if tag == "th":
                scope = f' scope="{"colgroup" if cs > 1 else "col"}"'
            parts.append(f"<{tag}{scope} rowspan='{rs}' colspan='{cs}'>{content}</{tag}>")
            col += cs
        else:
            # 检查是否在合并区域内 (非起始位置, 跳过)
            is_passive = False
            for (mr, mc), (mrs, mcs) in merge_map.items():
                if mr < global_row < mr + mrs and mc <= col < mc + mcs:
                    is_passive = True
                    break
                if mr == global_row and mc < col < mc + mcs:
                    is_passive = True
                    break
            if is_passive:
                col += 1
            else:
                cell_val = row[col] if col < len(row) else ""
                content = _cell_to_text(cell_val)
                parts.append(f"<{tag}>{content}</{tag}>")
                col += 1
    parts.append("</tr>")
    return "\n".join(parts)


def _cell_to_text(value) -> str:
    """将单元格值转为 HTML-safe 文本"""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        # 整数不显示小数点
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value)
    # 转义 HTML 特殊字符
    text = str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    # 换行 → <br>
    text = text.replace("\n", "<br>").replace("\r", "")
    return text


def _html_to_markdown_table(html: str) -> str:
    """
    将 HTML 表格转为 Markdown 表格。

    保留为 HTML (不转换) 的情况:
    1. 列数 > MAX_COLS_FOR_MARKDOWN_TABLE (太宽)
    2. 多层表头 (thead 中有 ≥ 2 行) — Markdown 无法表达
    3. 有合并单元格 (colspan > 1 或 rowspan > 1)
    """
    import re

    # ── 检测: 多层表头? ──
    thead_match = re.search(r'<thead>(.*?)</thead>', html, re.DOTALL | re.IGNORECASE)
    if thead_match:
        header_tr_count = len(re.findall(r'<tr>', thead_match.group(1), re.IGNORECASE))
        if header_tr_count >= 2:
            # ★ 多层表头 — Markdown 无法表达 → 保留 HTML
            return html

    # ── 检测: 合并单元格? ──
    has_colspan = bool(re.search(r'colspan\s*=\s*["\']?\s*(\d+)', html, re.IGNORECASE))
    has_rowspan = bool(re.search(r'rowspan\s*=\s*["\']?\s*(\d+)', html, re.IGNORECASE))
    if has_colspan or has_rowspan:
        for m in re.finditer(r'colspan\s*=\s*["\']?\s*(\d+)', html, re.IGNORECASE):
            if int(m.group(1)) > 1:
                return html  # 有跨列合并 → 保留 HTML
        for m in re.finditer(r'rowspan\s*=\s*["\']?\s*(\d+)', html, re.IGNORECASE):
            if int(m.group(1)) > 1:
                return html  # 有跨行合并 → 保留 HTML

    # ── 检测: 列数 ──
    first_tr = re.search(r'<tr>(.*?)</tr>', html, re.DOTALL)
    if not first_tr:
        return html

    col_count = len(re.findall(r'<(th|td)', first_tr.group(1), re.IGNORECASE))
    if col_count > MAX_COLS_FOR_MARKDOWN_TABLE:
        return html

    # ── 转换: HTML → Markdown 表格 (仅单层表头 + 无合并单元格) ──
    md_lines = []
    trs = re.findall(r'<tr>(.*?)</tr>', html, re.DOTALL | re.IGNORECASE)
    is_first = True

    for tr in trs:
        cells = re.findall(r'<(th|td)([^>]*)>(.*?)</\1>', tr, re.DOTALL | re.IGNORECASE)
        row_texts = []
        for _, _, content in cells:
            content = re.sub(r'<[^>]+>', '', content).strip()
            content = content.replace("|", "\\|")
            row_texts.append(content)

        if row_texts:
            md_lines.append("| " + " | ".join(row_texts) + " |")
            if is_first:
                md_lines.append("| " + " | ".join(["---"] * len(row_texts)) + " |")
                is_first = False

    return "\n".join(md_lines) if md_lines else html


# ── 主处理逻辑 ────────────────────────────────────────────

async def process_excel(
    file_bytes: bytes,
    file_name: str = "document.xlsx",
    dify_client=None,        # DifyWorkflowClient | None
    max_dify_rows: int = DIFY_MAX_ROWS_PER_SHEET,
    skip_dify_threshold: int = DIFY_SKIP_ROWS_THRESHOLD,
) -> ExcelResult:
    """
    处理 Excel 文件: 读取所有 Sheet → 转 HTML → Dify 优化 → Markdown 输出。

    Args:
        file_bytes: Excel 文件字节流
        file_name: 文件名 (用于日志)
        dify_client: Dify 客户端 (None 则跳过优化)
        max_dify_rows: 单 Sheet 超此值分批送 Dify
        skip_dify_threshold: 单 Sheet 超此值跳过 Dify

    Returns:
        ExcelResult 包含所有 Sheet 的处理结果和 Markdown
    """
    import openpyxl

    start_time = time.time()
    result = ExcelResult(file_name=file_name)

    file_size = len(file_bytes)
    use_read_only = file_size > MAX_FILE_SIZE_NORMAL_MODE

    logger.info(
        f"Excel 处理: {file_name} ({file_size / 1024:.0f} KB), "
        f"mode={'read_only' if use_read_only else 'normal'}"
    )

    # ── 读取 ──
    try:
        wb = openpyxl.load_workbook(
            BytesIO(file_bytes),
            read_only=use_read_only,
            data_only=True,       # 只读取值, 不读取公式
        )
    except Exception as e:
        result.errors.append(f"Failed to open Excel: {e}")
        return result

    sheet_names = wb.sheetnames
    logger.info(f"Excel: {len(sheet_names)} sheets → {sheet_names}")

    for sheet_name in sheet_names:
        try:
            ws = wb[sheet_name]

            # 读取数据
            if use_read_only:
                rows = _read_sheet_streaming(ws)
                merged = []
            else:
                rows, merged = _read_sheet_normal(ws)

            row_count = len(rows)
            col_count = max(len(r) for r in rows) if rows else 0

            logger.debug(
                f"  Sheet '{sheet_name}': {row_count} rows × {col_count} cols"
            )

            if row_count == 0:
                result.sheets.append(SheetResult(
                    name=sheet_name, row_count=0, col_count=0,
                    html="", optimized_html="", markdown=f"## {sheet_name}\n\n*(empty sheet)*\n",
                ))
                continue

            # 转为 HTML
            html = _rows_to_html(rows, merged)

            # ── Dify 优化 ──
            optimized_html = ""
            dify_enhanced = False

            if dify_client and dify_client.table_configured:
                if row_count > skip_dify_threshold:
                    logger.info(
                        f"  Sheet '{sheet_name}': {row_count} rows > "
                        f"{skip_dify_threshold} threshold → skipping Dify"
                    )
                elif row_count <= max_dify_rows:
                    # 单次发送
                    try:
                        dify_result = await dify_client.optimize_table(
                            table_html=html,
                            table_index=0,
                            page_number=0,
                            caption=sheet_name,
                        )
                        optimized_html = dify_result.optimized_html or html
                        dify_enhanced = True
                        result.dify_calls += 1
                    except Exception as e:
                        logger.warning(f"  Dify optimize failed for '{sheet_name}': {e}")
                        optimized_html = html
                else:
                    # 分块发送
                    optimized_html = await _optimize_table_chunked(
                        dify_client, rows, merged,
                        sheet_name, row_count, col_count,
                        max_dify_rows,
                    )
                    dify_enhanced = True
                    result.dify_calls += (row_count + MAX_ROWS_PER_CHUNK - 1) // MAX_ROWS_PER_CHUNK
            else:
                optimized_html = html

            # 生成 Markdown
            final_html = optimized_html or html
            md_table = _html_to_markdown_table(final_html)

            markdown = f"## {sheet_name}\n\n"
            if dify_enhanced:
                markdown += f"*({row_count} rows × {col_count} cols, Dify optimized)*\n\n"
            else:
                markdown += f"*({row_count} rows × {col_count} cols)*\n\n"
            markdown += md_table + "\n"

            result.sheets.append(SheetResult(
                name=sheet_name,
                row_count=row_count,
                col_count=col_count,
                html=html,
                optimized_html=optimized_html,
                markdown=markdown,
                dify_enhanced=dify_enhanced,
            ))

            result.total_rows += row_count

        except Exception as e:
            logger.error(f"  Sheet '{sheet_name}' failed: {e}")
            result.errors.append(f"Sheet '{sheet_name}': {e}")
            result.sheets.append(SheetResult(
                name=sheet_name,
                error=str(e),
                markdown=f"## {sheet_name}\n\n*Error: {e}*\n",
            ))

    wb.close()
    result.processing_time_s = round(time.time() - start_time, 2)

    logger.info(
        f"Excel 完成: {len(result.sheets)} sheets, {result.total_rows} total rows, "
        f"{result.dify_calls} Dify calls, {result.processing_time_s}s"
    )

    return result


async def _optimize_table_chunked(
    dify_client,
    rows: list[list],
    merged: list[tuple],
    sheet_name: str,
    total_rows: int,
    col_count: int,
    chunk_size: int = MAX_ROWS_PER_CHUNK,
) -> str:
    """
    将大表格分块送 Dify 优化, 合并结果。

    策略: 第一块含表头, 后续块不含表头行。
    """
    num_chunks = (total_rows + chunk_size - 1) // chunk_size
    logger.info(f"  Chunking '{sheet_name}': {total_rows} rows → {num_chunks} chunks")

    optimized_parts = []

    for chunk_idx in range(num_chunks):
        start_row = chunk_idx * chunk_size
        end_row = min(start_row + chunk_size, total_rows)

        if chunk_idx > 0:
            # 后续块: 包含表头行 + 当前数据行 (保持列结构一致)
            header_rows = [rows[0]]
            data_rows = rows[start_row:end_row]
            chunk_rows = header_rows + data_rows
            chunk_html = _rows_to_html(chunk_rows, None, max_rows=0, start_row=0)
        else:
            chunk_html = _rows_to_html(rows, None, max_rows=chunk_size, start_row=0)

        try:
            dify_result = await dify_client.optimize_table(
                table_html=chunk_html,
                table_index=chunk_idx,
                page_number=0,
                caption=f"{sheet_name} (part {chunk_idx + 1}/{num_chunks})",
            )
            opt = dify_result.optimized_html or chunk_html
            # 移除后续块中的重复表头行 (简单策略: 去掉第一个 <tr>)
            if chunk_idx > 0:
                import re
                opt = re.sub(r'<tr>.*?</tr>', '', opt, count=1, flags=re.DOTALL)
            optimized_parts.append(opt)
        except Exception as e:
            logger.warning(f"  Chunk {chunk_idx + 1}/{num_chunks} failed: {e}")
            optimized_parts.append(chunk_html)

        # 短暂等待避免 API 限流
        if chunk_idx < num_chunks - 1:
            await asyncio.sleep(0.5)

    # 合并
    if optimized_parts:
        merged_html = optimized_parts[0]
        for part in optimized_parts[1:]:
            # 在 </table> 之前插入
            merged_html = merged_html.replace("</table>", "") + "\n" + part
            # 移除内部 <table> 标签
            merged_html = merged_html.replace("<table>", "").replace("</table>", "")
            merged_html = "<table>\n" + merged_html + "\n</table>"

        return merged_html

    return ""


# ── Markdown 输出构建 ────────────────────────────────────

def build_excel_markdown(result: ExcelResult) -> str:
    """
    将 ExcelResult 中的所有 Sheet Markdown 合并为完整文档。
    """
    parts = [f"# {result.file_name}\n"]

    for sheet in result.sheets:
        parts.append(sheet.markdown)
        parts.append("")  # 空行分隔

    # 元数据
    parts.append("---\n")
    parts.append(f"*{len(result.sheets)} sheets, {result.total_rows} total rows, "
                 f"{result.dify_calls} Dify calls, {result.processing_time_s}s*\n")

    if result.errors:
        parts.append("\n### Errors\n")
        for err in result.errors:
            parts.append(f"- {err}\n")

    return "\n".join(parts)


# ── 便捷入口 ──────────────────────────────────────────────

async def parse_excel_to_markdown(
    file_path: str,
    output_dir: str,
    dify_client=None,
) -> dict:
    """
    一站式 Excel 解析入口。

    读取 Excel → 处理所有 Sheet → 输出 Markdown 到 output_dir。

    Returns:
        {"output_path": str, "result": ExcelResult}
    """
    path = Path(file_path)
    file_bytes = path.read_bytes()
    file_name = path.name

    result = await process_excel(
        file_bytes=file_bytes,
        file_name=file_name,
        dify_client=dify_client,
    )

    # 生成 Markdown
    markdown = build_excel_markdown(result)

    # 写入文件
    output_path = Path(output_dir) / f"{path.stem}.md"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(markdown, encoding="utf-8")

    # 同时输出每个 Sheet 的独立文件
    sheets_dir = Path(output_dir) / f"{path.stem}_sheets"
    sheets_dir.mkdir(parents=True, exist_ok=True)
    for sheet in result.sheets:
        sheet_path = sheets_dir / f"{_sanitize_name(sheet.name)}.md"
        sheet_path.write_text(sheet.markdown, encoding="utf-8")

    logger.info(f"Excel output: {output_path} + {sheets_dir}")

    return {
        "output_path": str(output_path),
        "sheets_dir": str(sheets_dir),
        "result": result,
    }


def _sanitize_name(name: str) -> str:
    """清理 Sheet 名称中的非法文件名字符"""
    import re
    return re.sub(r'[<>:"/\\|?*]', '_', name).strip() or "sheet"


# ── __main__ ──────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    async def _main():
        if len(sys.argv) < 3:
            print("Usage: python -m mineru.backend.rag.excel_processor <input.xlsx> <output_dir>")
            sys.exit(1)

        result = await parse_excel_to_markdown(sys.argv[1], sys.argv[2])
        print(f"Done: {result['output_path']}")

    asyncio.run(_main())
